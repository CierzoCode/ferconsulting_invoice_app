"""Extrae clientes, servicios y metadatos desde un .xlsm compatible con la plantilla original.
Uso:
    python scripts/extract_excel_data.py "Factura Ferconsulting 2026.xlsm"
"""
from __future__ import annotations
import json
import sys
from pathlib import Path
from openpyxl import load_workbook

BAD = {'#VALUE!', '#REF!', '#DIV/0!', '#NAME?', '#N/A', '#NULL!', '#NUM!'}

def clean(v):
    if v is None:
        return ''
    if isinstance(v, str):
        v = v.strip()
        if v in BAD:
            return ''
        return v[1:] if v.startswith("'") else v
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v

def main(path: str):
    out = Path(__file__).resolve().parents[1] / 'data'
    wb = load_workbook(path, data_only=True, keep_vba=True)
    clients = []
    ws = wb['CLIENTES']
    for r in range(2, ws.max_row + 1):
        name = clean(ws.cell(r, 2).value)
        if not name: continue
        clients.append({'id': len(clients)+1, 'source_key': f'CLIENTES:{r}', 'source_row': r, 'external_code': str(clean(ws.cell(r,1).value)), 'name': name, 'tax_id': str(clean(ws.cell(r,3).value)), 'address': str(clean(ws.cell(r,4).value)), 'postal_code': str(clean(ws.cell(r,5).value)), 'city': str(clean(ws.cell(r,6).value)), 'email': str(clean(ws.cell(r,7).value))})
    services = []
    ws = wb['SERVICIOS']
    for r in range(2, ws.max_row + 1):
        name = clean(ws.cell(r, 2).value)
        if not name: continue
        try: price = float(clean(ws.cell(r,4).value) or 0)
        except Exception: price = 0
        services.append({'id': len(services)+1, 'source_key': f'SERVICIOS:{r}', 'source_row': r, 'code': str(clean(ws.cell(r,1).value)), 'name': name, 'unit': str(clean(ws.cell(r,3).value)), 'unit_price': price, 'active': True})
    (out / 'clients.json').write_text(json.dumps(clients, ensure_ascii=False, indent=2), encoding='utf-8')
    (out / 'services.json').write_text(json.dumps(services, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f'Exportados {len(clients)} clientes y {len(services)} servicios.')

if __name__ == '__main__':
    if len(sys.argv) < 2:
        raise SystemExit('Indica la ruta del .xlsm')
    main(sys.argv[1])
